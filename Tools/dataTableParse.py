import os
import openpyxl
import json
import sys
import requests

def parseExcel(with_save=True, with_type=False):
    file_list = os.listdir('./Tables')
    total_result = {}
    for file in file_list:
        ext = os.path.splitext(file)[1]
        if '~' in file:
            continue
        if ext != '.xlsx':
            continue
        excel = openpyxl.load_workbook('./Tables/' +file, data_only=True, read_only=True)
        sheet = excel[excel.sheetnames[1]]
        print(file)
        print('row : {0}, col : {1}'.format(sheet.max_row, sheet.max_column))
        data = []
        type = []
        name = []
        file_name = os.path.splitext(file)[0]

        for col in range(1, sheet.max_column+1):
            type.append(sheet.cell(1, col).value)
            name.append(sheet.cell(4, col).value)

        current_row = 0
        for row in sheet.iter_rows(min_row=5):
            current_row = current_row + 1

            sub_data = {}
            if row[0].value == None:
                continue
            for col in range(1, sheet.max_column+1):
                value = row[col-1].value
                if with_type:
                    sub_data[name[col-1]+"_type"] = type[col-1]
                if type[col-1] == 'Int':
                    if value == '' or value == None: 
                        sub_data[name[col-1]] = 0
                    else:
                        sub_data[name[col-1]] = int(value)
                elif type[col-1] == 'Enum':
                    if value == '' or value == None: 
                        sub_data[name[col-1]] = 0
                    else:
                        sub_data[name[col-1]] = int(value)
                elif type[col-1] == 'Double':
                    if value == '' or value == None: 
                        sub_data[name[col-1]] = 0.0
                    else:
                        sub_data[name[col-1]] = round(float(value),5)
                elif type[col-1] == 'Float':
                    if value == '' or value == None: 
                        sub_data[name[col-1]] = 0.0
                    else:
                        sub_data[name[col-1]] = round(float(value),5)
                elif type[col-1] == 'Bool':
                    if value == '' or value == None: 
                        sub_data[name[col-1]] = False
                    else:
                        sub_data[name[col-1]] = bool(value)
                elif type[col-1] == 'String':
                        sub_data[name[col-1]] = value
                elif type[col-1] == 'Debug':
                    continue
                else:
                    print("{0} is invalid type!".format(type[col-1]))

            sub_data['code'] = row[0].value
            data.append(sub_data)

        total_result[file_name] = data
        result = {
            'data' : data
        }
        if with_save:
            result_file = open("../Content/Jsons/" +file_name + ".json", "w", encoding="UTF-8-sig")
            dumped_result = json.dumps(result, ensure_ascii=False)
            result_file.write(dumped_result)
            result_file.close()

    return total_result

if __name__ == "__main__":
    total_result = parseExcel(True, False)